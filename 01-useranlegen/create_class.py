import argparse
import random

from openpyxl import load_workbook


def read_excel(file: str):
    """
    reads an excel file and yields the lines
    :param file: excel file
    :return: lines of file
    """

    wb = load_workbook(file, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=1):
        name = row[0].value
        room_nr = row[1].value
        class_teacher = row[2].value

        if name is not None and room_nr is not None and class_teacher is not None:
            yield name, room_nr, class_teacher


def make_parser():
    '''
    creates parser to execute function on command line
    '''

    parser = argparse.ArgumentParser()
    parser.add_argument("filename")
    args = parser.parse_args()

    if args.filename:
        create_bash_file(args.filename)


def create_bash_file(filename):
    with open("class_script.sh", "w") as script:
        for script_line, script_line_password, usr_name in create_users(filename):
            script.write(script_line + script_line_password)

    with open("del_class_script.sh", "w") as del_script:
        for script_line in del_users(filename):
            del_script.write(script_line)

    with open("class_pw_list.txt", "w") as list:
        for script_line in user_pw_list(filename):
            list.write(script_line)



def create_users(filename: str):
    excel = read_excel(filename)
    excel.__next__()
    rand_chars = ["!", "%", "&", "(", ")", ",", ".", "_", "-", "=", "^", "#"]
    for line in excel:
        usr_name = "k"+line[0].lower()
        script_line = f"useradd -d /home/klassen/{usr_name} -c {usr_name} -m -g {line[0]} -G cdrom,plugdev,sambashare -s /bin/bash {usr_name}\n"
        script_line_password = f"echo {line[0]}:" + str(line[0]) + "\\" + random.choice(rand_chars) + str(line[1]) + "\\" + random.choice(rand_chars) + line[2] + "\\" + random.choice(rand_chars) + " | chpasswd\n\n"
        yield script_line, script_line_password, usr_name
    script_line = "useradd -d /home/lehrer/lehrer -c lehrer -m -g lehrer -G cdrom,plugdev,sambashare -s /bin/bash lehrer\n"
    script_line_password = f"echo lehrer:0{random.choice(rand_chars)}0{random.choice(rand_chars)}0{random.choice(rand_chars)} | chpasswd\n\n"
    yield script_line, script_line_password, usr_name
    script_line = "useradd -d /home/lehrer/seminar -c seminar -m -g seminar -Gcdrom,plugdev,sambashare -s /bin/bash seminar\n"
    script_line_password = f"echo seminar:1{random.choice(rand_chars)}1{random.choice(rand_chars)}1{random.choice(rand_chars)} | chpasswd\n\n"
    yield script_line, script_line_password, usr_name


def del_users(filename: str):
    excel = read_excel(filename)
    excel.__next__()
    for line in excel:
        usr_name = "k"+line[0].lower()
        script_line = f"userdel -r {usr_name}\n"
        yield script_line


def user_pw_list(filename: str):
    for script_line, script_line_password, usr_name in create_users(filename):
        pw = script_line_password.split()[1].split(':')[1].replace('\\', '')
        script_line = f"{usr_name}: {pw}\n"
        yield script_line


if __name__ == '__main__':
    make_parser()
    read_excel("./Klassenraeume_2023.xlsx")
