import argparse
import random
import string

import openpyxl
import unicodedata
from openpyxl import load_workbook

txtfile: bool = False


def read_excel(file: str):
    """
    reads an excel file and yields the lines
    :param file: excel file
    :return: lines of file
    """

    wb = load_workbook(file, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=1):
        firstname = row[0].value
        lastname = row[1].value
        type_name = row[2].value
        class_name = row[3].value

        if firstname is not None and lastname is not None and type is not None:
            yield firstname, lastname, type_name, class_name


def create_users(filename: str):
    excel = read_excel(filename)
    excel.__next__()
    usr_names = []
    rand_chars = ["!", "%", "&", "(", ")", ",", ".", "_", "-", "=", "^", "#"]
    for line in excel:
       # print(line[1])
        usr_name = line[1].lower().replace(" ", "_").replace('ß', 'ss').replace('ö', 'oe').replace('ä', 'ae').replace('ü', 'ue')
        usr_name = shave_marks(usr_name)
        i = 1
        while usr_name in usr_names:
            usr_name = shave_marks(line[1].lower()).replace(" ", "_")
            usr_name = usr_name+str(i)
            i += 1
        usr_names.append(usr_name)

        if line[2] == "teacher":
            group = "teacher"
            directory = f"/home/teachers/{usr_name}"

        else:
            group = line[3]
            directory = f"/home/students/{usr_name}"

        password = ''.join("\\" + random.choice(rand_chars) for _ in range(6))

        script_line = f"useradd -d {directory} -c {usr_name} -m -g {group} -G cdrom,plugdev,sambashare -s /bin/bash {usr_name}\n"
        password_line = f"echo {usr_name}:{password} | chpasswd\n\n"

        yield script_line, password_line, password, usr_name


def del_users(filename: str):
    excel = read_excel(filename)
    excel.__next__()
    for line in create_users(filename):
        usr_name = line[0].split()[-1]
        script_line = f"userdel -r {usr_name}\n"
        yield script_line


# def user_list(username: str, password: str, index: int):
    # excel = read_excel(filename)
    # excel.__next__()




def user_txt_list(filename: str):
    pass


def write_bash_file(filename: str):
    index: int = 2

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    with open("user_script.sh", "w") as script:
        for line, password_line, password, username in create_users(filename):
            sheet['A1'] = 'username'
            sheet['B1'] = 'password'

            sheet[f'A{index}'] = username
            sheet[f'B{index}'] = password

            script.write(line + password_line)
            index += 1
        workbook.save("user-password-list.xlsx")

    with open("del_user_script.sh", "w") as del_script:
        for script_line in del_users(filename):
            del_script.write(script_line)

    # if txtfile == False:
    #     user_list(filename)

#    else:
#        user_txt_list(filename)


def make_parser():
    '''
    creates parser to execute function on command line
    '''

    parser = argparse.ArgumentParser()
    parser.add_argument("filename")
    # parser.add_argument("-v", "--verbose", help="increase  output verbosity")
    # parser.add_argument("-x", "--excel", help="write created users and passwords in excel file")
    # parser.add_argument("-t", "--txt", help="write created users and passwords in txt file")
    args = parser.parse_args()

    # if args.verbose:
    #     print("verbosity turned on")

    # if args.excel:
    #     txtfile = False

    # if args.txt:
    #     txtfile = True

    if args.filename:
        write_bash_file(args.filename)




def shave_marks(txt):
    """Remove all diacritic marks"""
    norm_txt = unicodedata.normalize('NFD', txt)
    shaved = ''.join(c for c in norm_txt
                     if not unicodedata.combining(c))
    return unicodedata.normalize('NFC', shaved)


if __name__ == '__main__':
    make_parser()
    del_users("./Namen.xlsx")
