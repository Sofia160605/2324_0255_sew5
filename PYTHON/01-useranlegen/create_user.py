# Sofia Angerer

import argparse
import logging
import random
import sys
from logging.handlers import RotatingFileHandler
from plistlib import InvalidFileException

import openpyxl
import unicodedata
from openpyxl import load_workbook

txtfile: bool = False
logger = logging.getLogger("logger")

def read_excel(file: str):
    """
    reads an excel file and yields the lines
    :param file: excel file
    :return: lines of file
    """

    try:
        logger.debug("reading excel file")

        wb = load_workbook(file, read_only=True)
        ws = wb[wb.sheetnames[0]]
        existing_users = []
        for row in ws.iter_rows(min_row=1):
            firstname = row[0].value
            lastname = row[1].value
            type_name = row[2].value
            class_name = row[3].value

            if (firstname, lastname, type_name, class_name) in existing_users:
                print(f"User {firstname} {lastname} has already been created... you cannot create it again... what are you thinking...?")
            elif firstname is not None and lastname is not None and type is not None:
                existing_users.append((firstname, lastname, type_name, class_name))
                yield firstname, lastname, type_name, class_name
    except Exception:
        logger.error(f"file {file} not found :(")
        raise InvalidFileException


def create_users(filename: str):
    """
    Generates the necessary lines to create a create-user-script
    :param filename: excel file
    :return:

    >>>
    """
    excel = read_excel(filename)
    excel.__next__() # zeile überspringen
    usr_names = []
    rand_chars = ["!", "%", "&", "(", ")", ",", ".", "_", "-", "=", "^", "#"]
    for line in excel:
        lastname = line[1]
        usr_name = lastname.lower().replace(" ", "_").replace('ß', 'ss').replace('ö', 'oe').replace('ä', 'ae').replace(
            'ü', 'ue')
        usr_name = shave_marks(usr_name.lower()).replace(" ", "_")

        if lastname in usr_names:
            usr_name = handle_same_names(usr_names, lastname)

        usr_names.append(lastname)

        if line[2] == "teacher":
            group = "teacher"
            directory = f"/home/teachers/{usr_name}"

        else:
            group = line[3]
            directory = f"/home/students/{usr_name}"

        password = "a" + ''.join("\\" + random.choice(rand_chars) for _ in range(6))

        script_line = f"useradd -d {directory} -c {usr_name} -m -g {group} -G cdrom,plugdev,sambashare -s /bin/bash {usr_name}\n"
        password_line = f"echo {usr_name}:{password} | chpasswd\n\n"

        logger.debug(f"generating user {usr_name}")

        yield script_line, password_line, password, usr_name


def handle_same_names(lastnames: [], usr_name):
    """
    appends an increasing number to name if the name is already taken
    :param lastnames: list of last names already existing
    :param usr_name: last name that is already taken
    :return: user name with correct number appended

    >>> handle_same_names(["heinz", "heinz1", "huber", "huber", "huber"], "huber")
    'huber3'
    >>> handle_same_names(["heinz", "heinz", "huber", "huber", "huber"], "heinz")
    'heinz2'
    >>> handle_same_names(["heinz", "heinz", "huber", "huber", "huber"], "garfield")
    'garfield'
    """
    count = lastnames.count(usr_name)
    return f"{usr_name}{count}"


def del_users(filename: str):
    """
    Generates the necessary lines to create a del-user-script
    :param filename: excel file
    :return:
    """
    logger.debug("creating del_user_script")

    excel = read_excel(filename)
    excel.__next__()
    for line in create_users(filename):
        usr_name = line[0].split()[-1]
        script_line = f"userdel -r {usr_name}\n"
        yield script_line


def user_list(filename, data):
    """
    generates an excel file with the user names and the corresponding passwords
    :param filename: name of the file being generated
    :param data: user information
    """
    logger.debug("creating user-password-list")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    index = 2

    for line, password_line, password, username in data:
        sheet['A1'] = 'username'
        sheet['B1'] = 'password'

        sheet[f'A{index}'] = username
        sheet[f'B{index}'] = password.replace("\\", "")

        index += 1
    workbook.save(filename)


def user_txt_list(filename: str, data):
    """
    generates a text file with user names and the corresponding passwords
    :param filename: name of the file being generated
    :param data: user information
    """
    with open(filename, "w") as usr_pw_list:
        for line, password_line, password, username in data:
            password = password.replace("\\", "")
            usr_pw_list.write(f"{username}: {password}\n")


def write_bash_file(create_data, del_data):
    """
    uses the given lines to generate user-script and del-user-script
    :param create_data:
    :param del_data:
    :return:
    """
    with open("user_script.sh", "w") as script:
        for line, password_line, password, username in create_data:
            script.write(line + password_line)

    with open("del_user_script.sh", "w") as del_script:
        for script_line in del_data:
            del_script.write(script_line)


def make_parser():
    """
    creates parser to execute function on command line
    """

    create_logger()

    parser = argparse.ArgumentParser()
    parser.add_argument("filename")
    parser.add_argument("-v", "--verbose", help="increase output verbosity", action="store_true")
    parser.add_argument("-q", "--quiet", help="decrease output verbosity", action="store_true")
    parser.add_argument("-x", "--excel", help="write created users and passwords in excel file", action="store_true")
    parser.add_argument("-t", "--txt", help="write created users and passwords in txt file", action="store_true")
    args = parser.parse_args()

    if args.verbose:
        print("verbosity turned on")
        logger.setLevel(logging.DEBUG)

    if args.quiet:
        print("verbosity turned off")
        logger.setLevel(logging.NOTSET)

    users = list(create_users(args.filename))

    if args.txt:
        user_txt_list("outputfilename.txt", users)

    if args.excel:
        user_list("userlist.xlsx", users)

    if args.filename:
        write_bash_file(users, del_users(args.filename))


def create_logger():
    """
    creates logger necessary for logging
    """
    rotating_fh = RotatingFileHandler('create-user.log', maxBytes=10_000, backupCount=5)
    stream_h = logging.StreamHandler(sys.stdout)
    logger.addHandler(rotating_fh)
    logger.addHandler(stream_h)


def shave_marks(txt):
    """Remove all diacritic marks
    >>> shave_marks("Zöë")
    'Zoe'
    >>> shave_marks("î lövé sëw")
    'i love sew'
    """
    norm_txt = unicodedata.normalize('NFD', txt)
    shaved = ''.join(c for c in norm_txt
                     if not unicodedata.combining(c))
    return unicodedata.normalize('NFC', shaved)


if __name__ == '__main__':
    make_parser()
    del_users("./Namen.xlsx")
